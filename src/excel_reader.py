from datetime import date, datetime
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from guest import Guest
from reservation import Reservation


class ExcelReader:
    def read(self, filename: str) -> Reservation:
        file_path = Path(filename)

        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        workbook = load_workbook(
            filename=file_path,
            data_only=True,
            read_only=True,
        )

        if "Guests" not in workbook.sheetnames:
            raise ValueError("The workbook must contain a worksheet named 'Guests'.")

        worksheet = workbook["Guests"]
        rows = list(worksheet.iter_rows(values_only=True))

        if len(rows) < 5:
            raise ValueError("The Guests worksheet does not contain enough rows.")

        reservation_headers = self._clean_headers(rows[0])
        reservation_values = rows[1]
        reservation_record = dict(zip(reservation_headers, reservation_values))

        guest_headers = self._clean_headers(rows[4])
        guest_records: list[dict[str, Any]] = []

        for row_number, values in enumerate(rows[5:], start=6):
            record = dict(zip(guest_headers, values))

            first_name = self._optional_text(record.get("first_name"))
            document_number = self._optional_text(
                record.get("document_number")
            )

            if first_name is None and document_number is None:
                continue

            record["_row_number"] = row_number
            guest_records.append(record)



        if not guest_records:
            raise ValueError("No guest records were found.")

        reservation = Reservation(
            reservation_number=self._required_text(
                reservation_record,
                "reservation_number",
                row_number=2,
            ),
            accommodation=self._required_text(
                reservation_record,
                "codigoEstablecimiento",
                row_number=2,
            ),
            check_in=self._to_date(
                reservation_record.get("check_in"),
                "check_in",
                2,
            ),
            check_out=self._to_date(
                reservation_record.get("check_out"),
                "check_out",
                2,
            ),
            guests=[],
        )

        for record in guest_records:
            row_number = int(record["_row_number"])

            guest = Guest(
                first_name=self._required_text(
                    record,
                    "first_name",
                    row_number,
                ),
                first_surname=self._required_text(
                    record,
                    "first_surname",
                    row_number,
                ),
                second_surname=self._optional_text(record.get("second_surname")),
                date_of_birth=self._to_date(
                    record.get("date_of_birth"),
                    "date_of_birth",
                    row_number,
                ),
                document_type=self._required_text(
                    record,
                    "document_type",
                    row_number,
                ),
                document_number=self._required_text(
                    record,
                    "document_number",
                    row_number,
                ),
                security_code=self._optional_text(
                    record.get("soporteDocumento")
                )
                or "",
                nationality=self._required_text(
                    record,
                    "nationality",
                    row_number,
                ),
                phone_number=self._optional_text(
                    record.get("telefono")
                )
                or "",
                email=self._optional_text(
                    record.get("correo")
                )
                or "",
                address=self._required_text(
                    record,
                    "Direccion",
                    row_number,
                ),
            )
            reservation.guests.append(guest)

        return reservation
                
          

    @staticmethod
    def _clean_headers(values: tuple[Any, ...]) -> list[str]:
        headers: list[str] = []

        for value in values:
            if value is None:
                headers.append("")
            else:
                headers.append(str(value).strip())

        return headers

    @staticmethod
    def _optional_text(value: Any) -> str | None:
        if value is None:
            return None

        text = str(value).strip()
        return text or None

    def _required_text(
        self,
        record: dict[str, Any],
        field_name: str,
        row_number: int,
    ) -> str:
        value = self._optional_text(record.get(field_name))

        if value is None:
            raise ValueError(f"Row {row_number}: '{field_name}' is required.")

        return value

    @staticmethod
    def _to_date(
        value: Any,
        field_name: str,
        row_number: int,
    ) -> date:
        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, str):
            cleaned = value.strip()

            for date_format in (
                "%d/%m/%Y",
                "%Y-%m-%d",
                "%Y-%m-%dT%H:%M:%S",
            ):
                try:
                    return datetime.strptime(
                        cleaned,
                        date_format,
                    ).date()
                except ValueError:
                    continue

        raise ValueError(
            f"Row {row_number}: invalid date in " f"'{field_name}': {value!r}"
        )
